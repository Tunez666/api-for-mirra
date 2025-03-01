from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from ...models import Tablo


class Command(BaseCommand):
    help = 'Импорт данных из xlsx-файла в модель Tablo'

    def add_arguments(self, parser):
        parser.add_argument('path', type=str, help='Путь к xlsx-файлу')

    def handle(self, *args, **kwargs):
        fp = kwargs['path']
        wb = load_workbook(filename=fp)
        ws = wb[wb.sheetnames[0]]
        start = False
        id_u, name, oid_parameter, id_parameter, oid_interpretation = None, None, None, None, None
        id_interpretation, point_scale, valid, valid_author, valid_doi, org = None, None, None, None, None, None

        for row in ws.iter_rows(values_only=True):
            cells = [str(x).strip() if x is not None else '' for x in row]
            if not start:
                try:
                    id_u = cells.index("ID")
                    name = cells.index("NAME")
                    oid_parameter = cells.index("OID_PARAMETER")
                    id_parameter = cells.index("ID_PARAMETER")
                    oid_interpretation = cells.index("OID_INTERPRETATION")
                    id_interpretation = cells.index("ID_INTERPRETATION")
                    point_scale = cells.index("POINT_SCALE")
                    valid = cells.index("VALID")
                    valid_author = cells.index("VALID_AUTHOR")
                    valid_doi = cells.index("VALID_DOI")
                    org = cells.index("ORG")
                    start = True
                except ValueError as e:
                    self.stderr.write(self.style.ERROR(f"Ошибка: {e}"))
                    return
            else:
                clean_id_u = cells[id_u].strip()
                existing_record = Tablo.objects.filter(id_u=clean_id_u).first()
                if existing_record:
                    existing_record.name = cells[name].strip()
                    existing_record.oid_parameter = cells[oid_parameter].strip()
                    existing_record.id_parameter = cells[id_parameter].strip()
                    existing_record.oid_interpretation = cells[oid_interpretation].strip()
                    existing_record.id_interpretation = cells[id_interpretation].strip()
                    existing_record.point_scale = cells[point_scale].strip().lower() in ["true", "1", "yes"]
                    existing_record.valid = int(cells[valid]) if cells[valid].isdigit() else 0
                    existing_record.valid_author = cells[valid_author].strip()
                    existing_record.valid_doi = cells[valid_doi].strip()
                    existing_record.org = cells[org].strip()
                    existing_record.save()
                    self.stdout.write(self.style.WARNING(f'Обновлена существующая запись с ID: {clean_id_u}'))
                else:
                    new_record = Tablo(
                        id_u=clean_id_u,
                        name=cells[name].strip(),
                        oid_parameter=cells[oid_parameter].strip(),
                        id_parameter=cells[id_parameter].strip(),
                        oid_interpretation=cells[oid_interpretation].strip(),
                        id_interpretation=cells[id_interpretation].strip(),
                        point_scale=cells[point_scale].strip().lower() in ["true", "1", "yes"],
                        valid=int(cells[valid]) if cells[valid].isdigit() else 0,
                        valid_author=cells[valid_author].strip(),
                        valid_doi=cells[valid_doi].strip(),
                        org=cells[org].strip(),
                    )
                    new_record.save()
                    self.stdout.write(self.style.SUCCESS(f'Добавлена новая запись с ID: {clean_id_u}'))

        self.stdout.write(self.style.SUCCESS('Импорт завершен'))
